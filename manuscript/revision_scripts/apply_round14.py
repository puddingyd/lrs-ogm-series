#!/usr/bin/env python3
"""apply_round14.py — Re-score NCKUH_LRS_001 (Case 1 8p23.1 prox del)

Original 2A scoring (full +1.00 from "complete overlap of an established HI
region") was incorrect because the deletion does not include GATA4 — the
established haploinsufficient gene of the 8p23.1 microdeletion syndrome.

Strict Riggs 2020 re-scoring:
  1A (+0.00) protein-coding genes present
  2B (+0.00) partial overlap of established HI region; HI gene GATA4 NOT
             included in the deletion
  3A (+0.00) 11 protein-coding RefSeq genes (below 25-gene threshold)
  4A (+0.45) multiple unrelated probands carrying smaller deletions that
             are fully contained within the proband's deletion footprint
             have been reported in ClinVar (4 records) and DECIPHER (2
             patients) as Likely pathogenic / Pathogenic — Montenegro 2023
             and Ballarati 2011 are NOT used here because their published
             cases also include GATA4 and therefore do not match this
             non-cardiac proximal 8p23.1 phenotype.
  4F (+0.45) de novo origin confirmed by parental karyotyping at prenatal
             diagnosis (both parents normal karyotype)
  Total +0.90 -> Likely pathogenic.

Reads/writes:
  data/clinvar_submissions.xlsx       (in-place; row 6)
  data/breakpoint_coordinates.tsv     (in-place)
"""
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# ===========================================================================
# 1. ClinVar Excel — row 6 (NCKUH_LRS_001)
# ===========================================================================
XL = 'data/clinvar_submissions.xlsx'
COL_LOCAL_ID  = 1
COL_CLASSIF   = 34   # Germline classification
COL_SCORE     = 35   # Assertion score
COL_COMMENT   = 37   # Comment on classification

NEW_COMMENT = (
    'Classified per ACMG/ClinGen 2020 (Riggs et al.) for copy-number '
    'losses: 1A (protein-coding genes present, +0.00); 2B (partial '
    'overlap of an established haploinsufficient region — the 8p23.1 '
    'microdeletion syndrome region — but the established '
    'haploinsufficient gene GATA4 (ClinGen Dosage Sensitivity HI score '
    '3) is NOT included in this deletion; full 2A scoring therefore '
    'does not apply, +0.00); 3A (11 protein-coding RefSeq genes wholly '
    'or partially included — C8orf74, CLDN23, ERI1, MFHAS1, MSRA, '
    'PPP1R3B, PRAG1, PRSS55, RP1L1, SOX7, TNKS — below the 25-gene '
    'threshold, +0.00); 4A (multiple unrelated probands carrying '
    'smaller deletions fully contained within the proband\'s deletion '
    'footprint and classified as Likely pathogenic / Pathogenic in '
    'ClinVar [Variation IDs 4526777, 545352, 545372, 60347 — '
    'accessions VCV004526777.1, VCV000545352.1, VCV000545372.1, '
    'VCV000060347.2] and in DECIPHER [patients 332681 and 323624]; '
    '+0.45); 4F (de novo origin confirmed by parental karyotyping at '
    'prenatal diagnosis showing normal karyotypes in both parents; '
    '+0.45). Total score +0.90 -> Likely pathogenic.'
)

wb = openpyxl.load_workbook(XL, data_only=False)
ws = wb['Variant']

for r in range(6, 50):
    if ws.cell(r, COL_LOCAL_ID).value == 'NCKUH_LRS_001':
        ws.cell(r, COL_CLASSIF).value = 'Likely pathogenic'
        ws.cell(r, COL_SCORE).value   = '+0.90'
        ws.cell(r, COL_COMMENT).value = NEW_COMMENT
        print(f'  ClinVar row {r} NCKUH_LRS_001:')
        print(f'    classification: Pathogenic -> Likely pathogenic')
        print(f'    score:          +1.00 -> +0.90')
        print(f'    criteria:       2A (full overlap, +1.00)')
        print(f'                 -> 2B + 3A + 4A (ClinVar/DECIPHER LP/P subsets) + 4F (de novo)')
        break

wb.save(XL)
print(f'\n  Saved: {XL}\n')

# ===========================================================================
# 2. breakpoint_coordinates.tsv — case1_8p23.1_microdeletion
# ===========================================================================
TSV = 'data/breakpoint_coordinates.tsv'
with open(TSV) as f:
    lines = f.readlines()

new_lines = []
n_changed = 0
for line in lines:
    if line.startswith('1\tcase1_8p23.1_microdeletion\t'):
        line = line.replace('\tPathogenic\t+1.00', '\tLikely pathogenic\t+0.90')
        n_changed += 1
        print('  TSV: case1_8p23.1_microdeletion -> Likely pathogenic; score +0.90')
    new_lines.append(line)

with open(TSV, 'w') as f:
    f.writelines(new_lines)
print(f'  Saved: {TSV} ({n_changed} row updated)')

# Verify
import csv
print('\n=== Verification ===')
with open(TSV) as f:
    rdr = csv.DictReader(f, delimiter='\t')
    for row in rdr:
        if row['variant_id'] == 'case1_8p23.1_microdeletion':
            print(f'  variant_id:          {row["variant_id"]}')
            print(f'  acmg_classification: {row["acmg_classification"]}')
            print(f'  acmg_score:          {row["acmg_score"]}')
