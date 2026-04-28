#!/usr/bin/env python3
"""apply_round9.py
Round-9 patcher: address all errors raised by the ClinVar Excel validator
(see data/clinvar_submissions_fixed_validated.xlsx col 90 for the original
error list).

Fixes applied to data/clinvar_submissions.xlsx (Variant sheet, rows 6-10):

  1. Col 30 "Preferred condition name"  -> cleared
     (Condition is unambiguous from cols 28-29 type+value; ClinVar wants
     identifier OR name, not both — preferring the identifier.)

  2. Col 29 "Condition ID value" for HP-typed rows -> strip "HP:" prefix
     (validator parsed type "HP" + value "HP:0001263" as "HP:HP:0001263";
     value should be just "0001263" since the namespace prefix is in col 28.)

  3. Cols 6, 7, 8, 11, 16, 17, 18 -> cleared
     (Chromosome / Start / Stop / Variant type / Variant length / Copy
     number / Reference copy number — these conflict with HGVS in cols 4-5;
     ClinVar wants HGVS OR coordinates, not both. HGVS is the standard form
     for SVs and is preserved here.)

  4. Col 39 "Classification citations" -> cleared
     (Framework standard PMIDs — Riggs 2020 / Richards 2015 — must NOT be
     cited here per ClinVar; they go into "Assertion criteria" which is a
     submitter-profile / submission-portal setting, not in the spreadsheet.)

  5. Col 35 "Assertion score" -> cleared for all rows
     (Validator reports the submission portal's assertion method is "ACMG
     Guidelines, 2015" which is not point-based; the Riggs CNV scores are
     therefore inconsistent with the declared method. The numeric scores
     remain documented in col 37 "Comment on classification".)

Reads/writes: data/clinvar_submissions.xlsx (in place)
"""
import openpyxl
import warnings
warnings.filterwarnings('ignore')

XL = 'data/clinvar_submissions.xlsx'

COL = {
    'Local ID':                   1,
    'Reference sequence':         4,
    'HGVS':                       5,
    'Chromosome':                 6,
    'Start':                      7,
    'Stop':                       8,
    'Variant type':              11,
    'Variant length':            16,
    'Copy number':               17,
    'Reference copy number':     18,
    'Condition ID type':         28,
    'Condition ID value':        29,
    'Preferred condition name':  30,
    'Assertion score':           35,
    'Classification citations':  39,
}

CLEAR_FIELDS_FOR_HGVS  = ['Chromosome', 'Start', 'Stop', 'Variant type',
                          'Variant length', 'Copy number',
                          'Reference copy number']
CLEAR_ALWAYS_FIELDS    = ['Preferred condition name', 'Classification citations',
                          'Assertion score']

wb = openpyxl.load_workbook(XL, data_only=False)
ws = wb['Variant']

print('=== Round-9 fixes ===')
for r in range(6, 50):
    lid = ws.cell(r, COL['Local ID']).value
    if not lid:
        break

    print(f'\nRow {r}  {lid}:')

    # 1. Clear chromosome coordinates / variant type / CN fields (keep HGVS)
    for f in CLEAR_FIELDS_FOR_HGVS:
        ws.cell(r, COL[f]).value = None
    print(f'  cleared: {", ".join(CLEAR_FIELDS_FOR_HGVS)}')

    # 2. Clear "Preferred condition name", "Classification citations",
    #    "Assertion score"
    for f in CLEAR_ALWAYS_FIELDS:
        ws.cell(r, COL[f]).value = None
    print(f'  cleared: {", ".join(CLEAR_ALWAYS_FIELDS)}')

    # 3. Strip "HP:" prefix from Condition ID value when type is "HP"
    cid_type  = ws.cell(r, COL['Condition ID type']).value
    cid_value = ws.cell(r, COL['Condition ID value']).value
    if cid_type == 'HP' and cid_value and isinstance(cid_value, str) and cid_value.upper().startswith('HP:'):
        new_val = cid_value[3:]   # strip "HP:" / "hp:"
        ws.cell(r, COL['Condition ID value']).value = new_val
        print(f'  Condition ID value: "{cid_value}" -> "{new_val}"')

wb.save(XL)
print(f'\nSaved: {XL}')

# Read-back verification
wb2 = openpyxl.load_workbook(XL, data_only=False)
ws2 = wb2['Variant']
print('\n=== Verification ===')
for r in range(6, 11):
    print(f'\nRow {r}  {ws2.cell(r, 1).value}:')
    for label in ['Reference sequence', 'HGVS', 'Chromosome', 'Variant type',
                  'Condition ID type', 'Condition ID value',
                  'Preferred condition name', 'Assertion score',
                  'Classification citations']:
        v = ws2.cell(r, COL[label]).value
        v_disp = repr(v) if v is not None else '(empty)'
        print(f'  {label:30s} = {v_disp}')
