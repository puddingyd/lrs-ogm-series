#!/usr/bin/env python3
"""apply_round13.py
Round-13: drop PP4 from the two gene-disrupting inversions; downgrade
classifications.

  NCKUH_LRS_003  Case 2 AUTS2 inv      : PVS1 + PM6 + PP4  ->  PVS1 + PM6
                                          Pathogenic       ->  Likely pathogenic
  NCKUH_LRS_005  Case 3 CNTNAP2 inv    : PVS1_moderate + PM6 + PP4
                                                            ->  PVS1_moderate + PM6
                                          Likely pathogenic ->  Uncertain significance

Reads/writes:
  data/clinvar_submissions.xlsx       (in-place; rows 8 and 10)
  data/breakpoint_coordinates.tsv     (in-place; case2_AUTS2 and case3_CNTNAP2)
"""
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# ===========================================================================
# 1. ClinVar Excel — rows 8 (AUTS2) and 10 (CNTNAP2)
# ===========================================================================
XL = 'data/clinvar_submissions.xlsx'
COL_LOCAL_ID  = 1
COL_CLASSIF   = 34   # Germline classification
COL_COMMENT   = 37   # Comment on classification

NEW_AUTS2_COMMENT = (
    'Classified per ACMG/AMP 2015 (Richards et al.) for gene-disrupting '
    'copy-neutral structural variants. The proximal breakpoint of this '
    'inversion truncates AUTS2 in intron 2 and is predicted to result in '
    'loss-of-function. Criteria applied: PVS1 (predicted LoF in an '
    'established haploinsufficient gene; AUTS2 ClinGen Dosage Sensitivity '
    'haploinsufficiency score 3, sufficient evidence for autosomal-'
    'dominant haploinsufficiency); PM6 (de novo origin inferred from '
    'parental karyotyping rather than confirmed at base-pair resolution). '
    'Per the ACMG/AMP combinatorial rules, 1 Very Strong (PVS1) + '
    '1 Moderate (PM6) meets the criteria for Likely pathogenic.'
)

NEW_CNTNAP2_COMMENT = (
    'Classified per ACMG/AMP 2015 (Richards et al.) for gene-disrupting '
    'copy-neutral structural variants. The distal breakpoint of this '
    'inversion truncates CNTNAP2 and is predicted to result in '
    'loss-of-function. Criteria applied: PVS1_moderate (predicted LoF in '
    'CNTNAP2; ClinGen Dosage Sensitivity haploinsufficiency score 2, '
    'limited evidence for autosomal-dominant haploinsufficiency for '
    'ASD/language phenotypes in addition to the established AR Pitt-'
    'Hopkins-like-1 syndrome; PVS1 weight downgraded to Moderate per '
    'ClinGen Sequence Variant Interpretation Working Group guidance); '
    'PM6 (de novo origin inferred from parental karyotyping rather than '
    'confirmed at base-pair resolution). Per the ACMG/AMP combinatorial '
    'rules, 2 Moderate (PVS1_moderate + PM6) does not meet the '
    'classification thresholds for Likely pathogenic (which requires '
    '>= 3 Moderate, or 2 Moderate + >= 2 Supporting); the variant is '
    'therefore classified as Uncertain significance pending additional '
    'supporting evidence (e.g., functional studies, segregation, or '
    'recurrent observations in similarly affected unrelated probands).'
)

wb = openpyxl.load_workbook(XL, data_only=False)
ws = wb['Variant']
for r in range(6, 50):
    lid = ws.cell(r, COL_LOCAL_ID).value
    if not lid:
        break
    if lid == 'NCKUH_LRS_003':
        ws.cell(r, COL_CLASSIF).value = 'Likely pathogenic'
        ws.cell(r, COL_COMMENT).value = NEW_AUTS2_COMMENT
        print(f'  ClinVar row {r} (AUTS2 inv): Pathogenic -> Likely pathogenic; PP4 dropped')
    elif lid == 'NCKUH_LRS_005':
        ws.cell(r, COL_CLASSIF).value = 'Uncertain significance'
        ws.cell(r, COL_COMMENT).value = NEW_CNTNAP2_COMMENT
        print(f'  ClinVar row {r} (CNTNAP2 inv): Likely pathogenic -> Uncertain significance; PP4 dropped')

wb.save(XL)
print(f'  Saved: {XL}\n')

# ===========================================================================
# 2. breakpoint_coordinates.tsv — case2_AUTS2 and case3_CNTNAP2 rows
# ===========================================================================
TSV = 'data/breakpoint_coordinates.tsv'
with open(TSV) as f:
    lines = f.readlines()

new_lines = []
n_changed = 0
for line in lines:
    if line.startswith('2\tcase2_AUTS2_disrupting_inv\t'):
        # classification "Pathogenic" -> "Likely pathogenic"
        # score "NA (PVS1+PM6+PP4)" -> "NA (PVS1+PM6)"
        line = line.replace('\tPathogenic\tNA (PVS1+PM6+PP4)',
                            '\tLikely pathogenic\tNA (PVS1+PM6)')
        n_changed += 1
        print('  TSV: case2_AUTS2_disrupting_inv -> Likely pathogenic; criteria NA (PVS1+PM6)')
    elif line.startswith('3\tcase3_7q35_inversion_CNTNAP2\t'):
        # classification "Likely pathogenic" -> "VUS"
        # score "NA (PVS1_moderate+PM6+PP4)" -> "NA (PVS1_moderate+PM6)"
        line = line.replace('\tLikely pathogenic\tNA (PVS1_moderate+PM6+PP4)',
                            '\tVUS\tNA (PVS1_moderate+PM6)')
        n_changed += 1
        print('  TSV: case3_7q35_inversion_CNTNAP2 -> VUS; criteria NA (PVS1_moderate+PM6)')
    new_lines.append(line)

with open(TSV, 'w') as f:
    f.writelines(new_lines)
print(f'  Saved: {TSV} ({n_changed} rows updated)\n')

# Verify
import csv
print('=== Verification — TSV rows after edit ===')
with open(TSV) as f:
    rdr = csv.DictReader(f, delimiter='\t')
    for row in rdr:
        if row['variant_id'] in ('case2_AUTS2_disrupting_inv',
                                  'case3_7q35_inversion_CNTNAP2'):
            print(f'  {row["variant_id"]}: classification={row["acmg_classification"]:25s}  score={row["acmg_score"]}')
