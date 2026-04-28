#!/usr/bin/env python3
"""apply_round8.py
Round-8 patcher: per-variant ClinVar citation columns.

For each of the 5 P/LP/VUS variants, set:
  Col 39  Classification citations  (the ACMG framework reference used)
  Col 70  Evidence citations        (variant-/gene-specific supporting papers)

Reads/writes:
  data/clinvar_submissions.xlsx (in place)
"""
import openpyxl
import warnings
warnings.filterwarnings('ignore')

XL = 'data/clinvar_submissions.xlsx'

# ---------------------------------------------------------------------------
# Framework references (Classification citations, col 39)
# ---------------------------------------------------------------------------
RIGGS_2020   = 'PMID:32531644'   # ACMG/ClinGen 2020 CNV scoring (Riggs et al.)
RICHARDS_2015 = 'PMID:25741868'  # ACMG/AMP 2015 sequence-variant standards
                                 # (Richards et al.) — used for gene-disrupting
                                 # copy-neutral SVs (PVS1, PM6, PP4)

# ---------------------------------------------------------------------------
# Per-variant references (Evidence citations, col 70). All PMIDs verified
# from full-text citations supplied by the user.
# ---------------------------------------------------------------------------
PER_VARIANT = {
    # Local ID         (Classification citations,           Evidence citations)
    'NCKUH_LRS_001': (RIGGS_2020,
                      'PMID:36067875 | PMID:20969981'),
        # Montenegro 2023 (8p23.1 phenotype expansion) +
        # Ballarati 2011 (8p23.1 GATA4 dosage)
    'NCKUH_LRS_002': (RIGGS_2020,
                      ''),
        # No specific 8p22-only NDD evidence; intentionally left blank
    'NCKUH_LRS_003': (RICHARDS_2015,
                      'PMID:39953909'),
        # Loberti 2025 (AUTS2-related syndrome European cohort)
    'NCKUH_LRS_004': (RIGGS_2020,
                      'PMID:26064708 | PMID:28053794'),
        # Dilzell 2015 (7q33-q35 deletion) + Kale 2016 (7q33-q36.1 deletion)
    'NCKUH_LRS_005': (RICHARDS_2015,
                      'PMID:21082657 | PMID:19582487 | PMID:21310003 | PMID:36253443'),
        # Sehested 2010 (CNTNAP2 7q34-q36.2 deletion) +
        # Poot 2010 (CNTNAP2 disruption + ASD) +
        # Whitehouse 2011 (CNTNAP2 language) +
        # Jang 2023 (Cntnap2 molecular networks)
}

COL_LOCAL_ID    = 1
COL_CLASSIF_CIT = 39   # "Classification citations"
COL_EVID_CIT    = 70   # "Evidence citations"

wb = openpyxl.load_workbook(XL, data_only=False)
ws = wb['Variant']

print('=== Round-8 patches ===')
for r in range(6, 50):
    lid = ws.cell(r, COL_LOCAL_ID).value
    if not lid:
        break
    if lid in PER_VARIANT:
        clf_cit, evid_cit = PER_VARIANT[lid]
        old_clf = ws.cell(r, COL_CLASSIF_CIT).value
        ws.cell(r, COL_CLASSIF_CIT).value = clf_cit
        ws.cell(r, COL_EVID_CIT).value    = evid_cit if evid_cit else None
        print(f'  Row {r}  {lid}:')
        print(f'    Classification citations: {old_clf!r:55s} -> {clf_cit!r}')
        print(f'    Evidence citations:        (empty)                                                 -> {evid_cit!r}')

wb.save(XL)
print(f'\nSaved: {XL}')

# Verify
wb2 = openpyxl.load_workbook(XL, data_only=False)
ws2 = wb2['Variant']
print('\n=== Read-back ===')
for r in range(6, 11):
    print(f'  Row {r}  {ws2.cell(r,1).value:14s}  '
          f'cls_cit={ws2.cell(r,COL_CLASSIF_CIT).value:20s}  '
          f'evid_cit={ws2.cell(r,COL_EVID_CIT).value or "(empty)"}')
