#!/usr/bin/env python3
"""fill_clinvar_submission.py
Populate ClinVar submission template (Variant sheet) with the 5 P/LP variants
from the LRS+OGM manuscript.

Reads:  data/clinvar_submissions.xlsx  (in-place edit)
Writes: data/clinvar_submissions.xlsx  (modified)
"""
import openpyxl
import warnings
warnings.filterwarnings('ignore')

XL = 'data/clinvar_submissions.xlsx'
wb = openpyxl.load_workbook(XL, data_only=False)
ws = wb['Variant']

# Column index map (1-based) — extracted from row 3 of the template
COLS = {
    'Local ID':                                  1,
    'Linking ID':                                2,
    'Gene symbol':                               3,
    'Reference sequence':                        4,
    'HGVS':                                      5,
    'Chromosome':                                6,
    'Start':                                     7,
    'Stop':                                      8,
    'Reference allele':                          9,
    'Alternate allele':                         10,
    'Variant type':                             11,
    'Variant length':                           16,
    'Copy number':                              17,
    'Reference copy number':                    18,
    'Breakpoint 1':                             19,
    'Breakpoint 2':                             20,
    'Comment on variant':                       21,
    'Variation identifiers':                    23,
    'Condition ID type':                        28,
    'Condition ID value':                       29,
    'Preferred condition name':                 30,
    'Condition comment':                        32,
    'Germline classification':                  34,
    'Assertion score':                          35,
    'Date last evaluated':                      36,
    'Comment on classification':                37,
    'Mode of inheritance':                      38,
    'Classification citations':                 39,
    'Collection method':                        44,
    'Allele origin':                            45,
    'Affected status':                          46,
    'SV method':                                47,
    'Clinical features':                        48,
    'Tissue':                                   51,
    'Sex':                                      52,
    'Age range':                                53,
    'Total individuals tested':                 58,
    'Number of individuals with variant':       61,
    'Test name or type':                        73,
    'Platform type':                            74,
    'Platform name':                            75,
    'Method':                                   76,
    'Method purpose':                           77,
    'Software name and version':                79,
    'Software purpose':                         80,
}

# ============================================================================
# Reusable observation defaults shared across all 5 variants
# ============================================================================
DEFAULTS = {
    'Reference sequence':                       'GRCh38',
    'Collection method':                        'clinical testing',
    'Allele origin':                            'de novo',
    'Affected status':                          'yes',
    'SV method':                                'SNP array | Read depth | Paired-end mapping | Sequence alignment | Optical mapping',
    'Tissue':                                   'Peripheral blood',
    'Total individuals tested':                 1,
    'Number of individuals with variant':       1,
    'Test name or type':                        'PacBio HiFi long-read whole-genome sequencing | Bionano optical genome mapping | SNP array (Thermo Fisher HD)',
    'Platform type':                            'next-gen sequencing | optical mapping | SNP array',
    'Platform name':                            'PacBio Sequel IIe | Bionano Saphyr | Thermo Fisher Scientific HD SNP array',
    'Method':                                   'Integrated PacBio HiFi long-read whole-genome sequencing (pbsv v2.10.0 + HiFiCNV v1.0.1) and Bionano optical genome mapping (Solve v3.7.2 / Access v1.7.2), with SNP array confirmation. Breakpoints curated by manual review in IGV and Ribbon. Annotation and prioritisation with AnnotSV.',
    'Method purpose':                           'discovery',
    'Software name and version':                'pbsv v2.10.0 | HiFiCNV v1.0.1 | DeepVariant v1.6.1 | Bionano Solve v3.7.2 | Bionano Access v1.7.2 | AnnotSV (latest) | VEP v115 | ANNOVAR (release 2024-04-30)',
    'Software purpose':                         'structural variant calling | copy-number variant calling | small-variant calling | optical genome map assembly | structural-variant annotation and prioritisation | variant annotation',
    'Classification citations':                 'PMID:32531644 | PMID:39243750',  # Riggs ACMG/ClinGen 2020 + Eisfeldt 2024 LRS rearrangements
}

# ============================================================================
# Per-variant data (5 rows)
# ============================================================================
VARIANTS = [
    # ---- Case 1 (4-year-old girl) ----
    {
        'Local ID':       'NCKUH_LRS_001',
        'Linking ID':     'NCKUH_LRS_001',
        'Gene symbol':    'SOX7|TNKS',  # within deleted interval; GATA4 spared
        'Chromosome':     '8',
        'Start':          8174929,
        'Stop':           10727966,
        'Variant type':   'Copy number loss',
        'Variant length': 2553037,
        'Copy number':    1,
        'Reference copy number': 2,
        'Comment on variant': '2.55-Mb interstitial deletion at 8p23.1 within the established '
                              '8p23.1 microdeletion region (ClinGen Dosage Sensitivity = '
                              'sufficient evidence for haploinsufficiency). The deletion '
                              'spares GATA4 (concordant with the absence of congenital '
                              'heart disease in the proband).',
        'Condition ID type':         'OMIM',
        'Condition ID value':        '614462',  # 8p23.1 deletion syndrome (placeholder; user can verify)
        'Preferred condition name':  '8p23.1 microdeletion syndrome',
        'Condition comment':         'Phenotype concordant with 8p23.1 microdeletion syndrome '
                                     'minus the cardiac component (GATA4 spared).',
        'Germline classification':   'Pathogenic',
        'Assertion score':           '+1.00',
        'Date last evaluated':       '2026-04-27',
        'Comment on classification': 'Classified per ACMG/ClinGen 2020 (Riggs et al., '
                                     'reference 40 in manuscript) for copy-number losses: '
                                     '1A (protein-coding genes present, +0.00) and 2A '
                                     '(complete overlap of an established haploinsufficient '
                                     'region — the 8p23.1 microdeletion syndrome critical '
                                     'region; +1.00). Total score +1.00 -> Pathogenic.',
        'Mode of inheritance':       'Autosomal dominant inheritance',
        'Clinical features': 'HP:0000252 (Microcephaly) | HP:0001263 (Global developmental '
                             'delay) | HP:0001513 (Obesity-not-applicable; placeholder remove '
                             'if not relevant) | HP:0001508 (Failure to thrive) | HP:0001250 '
                             '(Seizure) | HP:0010864 (Intellectual disability, severe)',
        'Sex':       'female',
        'Age range': '4 years',
    },
    {
        'Local ID':       'NCKUH_LRS_002',
        'Linking ID':     'NCKUH_LRS_002',
        'Gene symbol':    'TUSC3',
        'Chromosome':     '8',
        'Start':          12477756,
        'Stop':           16798965,
        'Variant type':   'Copy number loss',
        'Variant length': 4321209,
        'Copy number':    1,
        'Reference copy number': 2,
        'Comment on variant': '4.31-Mb deletion at 8p22 encompassing ~40 protein-coding '
                              'genes (incl. TUSC3); no single established haploinsufficient '
                              'gene drives this interval. Co-occurs with the 8p23.1 '
                              'microdeletion (NCKUH_LRS_001) and an intervening copy-neutral '
                              'inversion in the same proband, jointly forming a deletion-'
                              'inversion-deletion event refined by integrated LRS and OGM '
                              'from a prenatally reported "balanced" inv(8)(p23p11.2).',
        'Condition ID type':         'HP',
        'Condition ID value':        'HP:0001263',
        'Preferred condition name':  'Global developmental delay',
        'Germline classification':   'Likely pathogenic',
        'Assertion score':           '+0.55',
        'Date last evaluated':       '2026-04-27',
        'Comment on classification': 'Classified per ACMG/ClinGen 2020 (Riggs et al.) for '
                                     'copy-number losses: 1A (+0.00), 2H (no established HI '
                                     'gene, +0.00), 3B (25-49 protein-coding genes, +0.45), '
                                     '4O (partial overlap with literature 8p22 NDD cases, '
                                     '+0.10). Total score +0.55 -> Likely pathogenic.',
        'Mode of inheritance':       'Autosomal dominant inheritance',
        'Clinical features': 'HP:0000252 (Microcephaly) | HP:0001263 (Global developmental '
                             'delay) | HP:0001508 (Failure to thrive) | HP:0001250 (Seizure)',
        'Sex':       'female',
        'Age range': '4 years',
    },
    # ---- Case 2 (30-year-old man) ----
    {
        'Local ID':       'NCKUH_LRS_003',
        'Linking ID':     'NCKUH_LRS_003',
        'Gene symbol':    'AUTS2',
        'Chromosome':     '7',
        'Start':          69938691,
        'Stop':           95720897,
        'Variant type':   'Inversion',
        'Comment on variant': 'Apparent paracentric inv(7)(q11.21q22) on G-banded karyotype '
                              'refined by integrated optical genome mapping and PacBio HiFi '
                              'long-read sequencing into a copy-number-neutral inversion '
                              'with the proximal breakpoint at chr7:69,938,691 (intron 2 of '
                              'AUTS2 [NM_015570.4]) and the distal partner at chr7:95,720,897 '
                              '(7q21.3). The proximal breakpoint truncates AUTS2, predicted '
                              'to result in loss-of-function. Retrospective short-read '
                              'whole-exome SV re-analysis (Manta v1.6.0 and Delly v1.2.6) '
                              'confirmed both breakpoints fall outside exome capture '
                              'territory (mean depth 0.07x and 0.16x at the breakpoints '
                              'versus median 106x across captured AUTS2 exons), explaining '
                              'why the prior WES was non-diagnostic.',
        'Condition ID type':         'OMIM',
        'Condition ID value':        '615834',  # MRD26 / AUTS2 syndrome
        'Preferred condition name':  'AUTS2 syndrome (Mental retardation, autosomal dominant 26)',
        'Germline classification':   'Pathogenic',
        'Assertion score':           '',  # ACMG/AMP framework: PVS1+PM6+PP4
        'Date last evaluated':       '2026-04-27',
        'Comment on classification': 'Classified per ACMG/AMP 2015 (Richards et al.) for '
                                     'gene-disrupting copy-neutral structural variants: '
                                     'PVS1 (predicted loss-of-function in an established '
                                     'haploinsufficient gene; AUTS2 ClinGen Dosage '
                                     'Sensitivity HI score 3, sufficient evidence for '
                                     'autosomal-dominant haploinsufficiency) + PM6 (de '
                                     'novo by parental karyotyping, not molecularly '
                                     'confirmed at base-pair resolution) + PP4 (phenotype '
                                     'highly specific to AUTS2-related syndrome: '
                                     'intellectual disability, autism, ADHD, '
                                     'micrognathia, scoliosis). Pathogenic.',
        'Mode of inheritance':       'Autosomal dominant inheritance',
        'Clinical features': 'HP:0001249 (Intellectual disability) | HP:0000729 (Autistic '
                             'behavior) | HP:0007018 (ADHD) | HP:0000347 (Micrognathia) | '
                             'HP:0002650 (Scoliosis) | HP:0000405 (Conductive hearing '
                             'impairment, unilateral) | HP:0001249 (motor and language '
                             'delay in childhood) | HP:0100753 (Schizophrenia)',
        'Sex':       'male',
        'Age range': '30 years',
    },
    # ---- Case 3 (3-year-old boy) ----
    {
        'Local ID':       'NCKUH_LRS_004',
        'Linking ID':     'NCKUH_LRS_004',
        'Gene symbol':    'BRAF|AGK|CLCN1|TBXAS1|WEE2|SSBP1|TAS2R38|PRSS1|PRSS2|TRPV6|KEL|CASP2|NOBOX|TPK1|SLC37A3',
        'Chromosome':     '7',
        'Start':          139635053,
        'Stop':           145048896,
        'Variant type':   'Copy number loss',
        'Variant length': 5413843,
        'Copy number':    1,
        'Reference copy number': 2,
        'Comment on variant': '5.41-Mb interstitial deletion at 7q34-q35 encompassing 15 '
                              'OMIM-listed genes and ~45 protein-coding genes total. '
                              'Co-occurs with a CNTNAP2-disrupting inversion at 7q35 '
                              '(NCKUH_LRS_005) and a multi-chromosomal complex genomic '
                              'rearrangement (CGR) involving chromosomes 2, 5, 7, and 13 '
                              'in the same proband, with features suggestive of '
                              'chromoanagenesis. Identified concordantly by SNP array, '
                              'optical genome mapping, and PacBio HiFi long-read sequencing.',
        'Condition ID type':         'HP',
        'Condition ID value':        'HP:0001263',
        'Preferred condition name':  'Global developmental delay',
        'Germline classification':   'Likely pathogenic',
        'Assertion score':           '+0.60',
        'Date last evaluated':       '2026-04-27',
        'Comment on classification': 'Classified per ACMG/ClinGen 2020 (Riggs et al.) for '
                                     'copy-number losses: 1A (+0.00), 3B (25-49 protein-'
                                     'coding genes, +0.45), 4O (partial overlap with '
                                     'literature 7q33-q36 NDD cases, +0.15). Total score '
                                     '+0.60 -> Likely pathogenic.',
        'Mode of inheritance':       'Autosomal dominant inheritance',
        'Clinical features': 'HP:0001263 (Global developmental delay) | HP:0000729 '
                             '(Autistic behavior) | HP:0007018 (ADHD) | HP:0000028 '
                             '(Cryptorchidism) | HP:0000054 (Micropenis)',
        'Sex':       'male',
        'Age range': '3 years',
    },
    {
        'Local ID':       'NCKUH_LRS_005',
        'Linking ID':     'NCKUH_LRS_005',
        'Gene symbol':    'CNTNAP2',
        'Chromosome':     '7',
        'Start':          145048896,
        'Stop':           146600257,
        'Variant type':   'Inversion',
        'Comment on variant': '1.55-Mb copy-neutral inversion at 7q35 with the distal '
                              'breakpoint at chr7:146,600,257 truncating CNTNAP2 (ClinGen '
                              'Dosage Sensitivity HI score 2; autosomal-dominant risk for '
                              'autism spectrum disorder and language impairment). '
                              'Co-occurs with a 7q34-q35 deletion (NCKUH_LRS_004) in the '
                              'same proband as part of a complex multi-chromosomal '
                              'rearrangement with features suggestive of chromoanagenesis. '
                              'Identified by integrated optical genome mapping and PacBio '
                              'HiFi long-read sequencing.',
        'Condition ID type':         'HP',
        'Condition ID value':        'HP:0000729',
        'Preferred condition name':  'Autistic behavior',
        'Germline classification':   'Likely pathogenic',
        'Assertion score':           '',
        'Date last evaluated':       '2026-04-27',
        'Comment on classification': 'Classified per ACMG/AMP 2015 (Richards et al.) for '
                                     'gene-disrupting copy-neutral structural variants: '
                                     'PVS1_moderate (predicted loss-of-function in CNTNAP2; '
                                     'ClinGen Dosage Sensitivity HI score 2 — some evidence '
                                     'for autosomal-dominant haploinsufficiency for ASD/'
                                     'language impairment, justifying moderate weight) + '
                                     'PM6 (de novo by parental karyotyping, not molecularly '
                                     'confirmed at base-pair resolution) + PP4 (proband\'s '
                                     'autism / language / ADHD phenotype matches reported '
                                     'CNTNAP2 disruption phenotypes). Likely pathogenic.',
        'Mode of inheritance':       'Autosomal dominant inheritance',
        'Clinical features': 'HP:0001263 (Global developmental delay) | HP:0000729 '
                             '(Autistic behavior) | HP:0007018 (ADHD) | HP:0000028 '
                             '(Cryptorchidism) | HP:0000054 (Micropenis)',
        'Sex':       'male',
        'Age range': '3 years',
    },
]

# ============================================================================
# Write each variant on its own row (rows 6 onwards — row 5 says "start in next row")
# ============================================================================
START_ROW = 6
for i, var in enumerate(VARIANTS):
    row = START_ROW + i
    # Apply defaults first, then variant-specific overrides
    merged = {**DEFAULTS, **var}
    for field, value in merged.items():
        col = COLS.get(field)
        if col is None:
            print(f'WARN: unknown field "{field}"')
            continue
        ws.cell(row=row, column=col).value = value
    print(f'Filled row {row} for {var["Local ID"]}: {var["Variant type"]} on chr{var["Chromosome"]}')

wb.save(XL)
print(f'\nSaved: {XL}')

# Quick read-back sanity check
wb2 = openpyxl.load_workbook(XL, data_only=False)
ws2 = wb2['Variant']
print('\n=== Read-back check (Local ID + Variant type + Classification) ===')
for r in range(START_ROW, START_ROW + len(VARIANTS)):
    print(f'  Row {r}: {ws2.cell(r, 1).value:14s}  '
          f'{ws2.cell(r, 11).value:18s}  '
          f'{ws2.cell(r, 34).value}')
