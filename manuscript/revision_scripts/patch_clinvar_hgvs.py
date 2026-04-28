#!/usr/bin/env python3
"""patch_clinvar_hgvs.py
Add HGVS expressions to the 5 ClinVar variant rows and fix the 8p23.1
condition (no OMIM number exists -> HPO HP:0001263).
"""
import openpyxl
import warnings
warnings.filterwarnings('ignore')

XL = 'data/clinvar_submissions.xlsx'

# RefSeq accessions for GRCh38 chromosomes used in our variants
REFSEQ = {
    '7': 'NC_000007.14',
    '8': 'NC_000008.11',
}

# (Local ID, Reference sequence, HGVS expression — g. portion)
PATCHES = [
    ('NCKUH_LRS_001', REFSEQ['8'], 'g.8174929_10727966del'),
    ('NCKUH_LRS_002', REFSEQ['8'], 'g.12477756_16798965del'),
    ('NCKUH_LRS_003', REFSEQ['7'], 'g.69938691_95720897inv'),
    ('NCKUH_LRS_004', REFSEQ['7'], 'g.139635053_145048896del'),
    ('NCKUH_LRS_005', REFSEQ['7'], 'g.145048896_146600257inv'),
]

# Column indices
COL_LOCAL_ID = 1
COL_REFSEQ   = 4
COL_HGVS     = 5
COL_COND_ID_TYPE  = 28
COL_COND_ID_VALUE = 29
COL_COND_NAME     = 30

wb = openpyxl.load_workbook(XL, data_only=False)
ws = wb['Variant']

# Walk rows 6+ and apply patches by Local ID
for r in range(6, 50):
    local_id = ws.cell(r, COL_LOCAL_ID).value
    if not local_id:
        break
    for (lid, refseq, hgvs) in PATCHES:
        if local_id == lid:
            ws.cell(r, COL_REFSEQ).value = refseq
            ws.cell(r, COL_HGVS).value = hgvs
            print(f'  Row {r}  {lid}: refseq={refseq}, hgvs={hgvs}')
            break

# Fix Case 1 8p23.1 condition: no OMIM exists for this microdeletion syndrome
# -> use HPO HP:0001263 Global developmental delay (matches the other Case 1 row)
for r in range(6, 50):
    if ws.cell(r, COL_LOCAL_ID).value == 'NCKUH_LRS_001':
        ws.cell(r, COL_COND_ID_TYPE).value  = 'HP'
        ws.cell(r, COL_COND_ID_VALUE).value = 'HP:0001263'
        ws.cell(r, COL_COND_NAME).value     = 'Global developmental delay'
        print(f'  Row {r}  NCKUH_LRS_001 condition: OMIM 614462 -> HP:0001263 (no OMIM exists for 8p23.1 microdeletion syndrome)')
        break

wb.save(XL)
print(f'\nSaved: {XL}')

# Verify
wb2 = openpyxl.load_workbook(XL, data_only=False)
ws2 = wb2['Variant']
print('\n=== Read-back: cols 1, 4, 5, 28, 29, 30 for rows 6-10 ===')
for r in range(6, 11):
    print(f'  Row {r}: '
          f'{str(ws2.cell(r,1).value):16s} | '
          f'{str(ws2.cell(r,4).value):14s} | '
          f'{str(ws2.cell(r,5).value):32s} | '
          f'{str(ws2.cell(r,28).value):6s} | '
          f'{str(ws2.cell(r,29).value):12s} | '
          f'{ws2.cell(r,30).value}')
